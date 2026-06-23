#!/usr/bin/env python3
"""Erzeugt die zentrale Excel-Vorlage für Private-Access-Importe."""
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:
    raise SystemExit(
        'openpyxl fehlt. Installieren: pip install -r import/requirements.txt'
    ) from exc

HEADERS = [
    'application_name',
    'connector_group',
    'entra_group_name',
    'target_type',
    'fqdn',
    'ip_address',
    'cidr_range',
    'ip_range_start',
    'ip_range_end',
    'protocol',
    'ports',
    'change_reference',
    'workload',
    'source_reference',
    'review_status',
    'review_comment',
    'import_status',
    'generated_yaml',
    'imported_at',
    'import_note',
]

EXAMPLES = [
    [
        'EIT_PrivateAccess_NST_JumpHosts',
        'Nordbayern Frontend',
        'EIT_Intune_EIT_cfg Private Access PILOT',
        'fqdn',
        'host01.nb.edeka.net',
        '',
        '',
        '',
        '',
        'both',
        '3389',
        'PILOT-JUMPHOSTS-001',
        'jumphosts',
        'Beispiel FQDN',
        'approved',
        '',
        'offen',
        '',
        '',
        '',
    ],
    [
        '',
        '',
        '',
        'fqdn',
        'host02.nb.edeka.net',
        '',
        '',
        '',
        '',
        'both',
        '3389',
        '',
        '',
        'Connector/App von Zeile oben übernehmen (leer lassen)',
        'approved',
        '',
        'offen',
        '',
        '',
        '',
    ],
    [
        'EIT_PrivateAccess_NST_DCs',
        'Nordbayern Frontend',
        'EIT_Intune_EIT_cfg Private Access PILOT',
        'fqdn',
        'dc01.nb.edeka.net',
        '',
        '',
        '',
        '',
        'both',
        '389,636,88,445',
        'CHG-DC-IMPORT-001',
        'domain-controllers',
        'Mehrere Ports kommagetrennt',
        'pending',
        'Ports mit Identity abstimmen',
        'offen',
        '',
        '',
        '',
    ],
]

OUTPUT_COLUMNS = {'import_status', 'generated_yaml', 'imported_at', 'import_note'}


def main():
    repo = Path(__file__).resolve().parents[2]
    out = repo / 'import' / 'templates' / 'private-access-import-template.xlsx'
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Import'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(color='FFFFFF', bold=True)
    output_fill = PatternFill('solid', fgColor='E7E6E6')

    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill if name not in OUTPUT_COLUMNS else output_fill

    for row_idx, example in enumerate(EXAMPLES, start=2):
        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{chr(64 + len(HEADERS))}1'

    for col, width in enumerate([34, 24, 38, 14, 28, 16, 16, 14, 14, 10, 14, 22, 18, 36, 14, 24, 14, 34, 22, 36], start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    review_dv = DataValidation(type='list', formula1='"approved,pending,rejected,needs_clarification"', allow_blank=True)
    review_dv.add(f'O2:O5000')
    ws.add_data_validation(review_dv)

    import_dv = DataValidation(type='list', formula1='"offen,übernommen,übersprungen,fehler"', allow_blank=True)
    import_dv.add(f'Q2:Q5000')
    ws.add_data_validation(import_dv)

    target_dv = DataValidation(type='list', formula1='"fqdn,ipAddress,ipRangeCidr,ipRange,dnsSuffix"', allow_blank=True)
    target_dv.add(f'D2:D5000')
    ws.add_data_validation(target_dv)

    proto_dv = DataValidation(type='list', formula1='"tcp,udp,both"', allow_blank=True)
    proto_dv.add(f'J2:J5000')
    ws.add_data_validation(proto_dv)

    wb.save(out)
    print(f'Vorlage erstellt: {out}')


if __name__ == '__main__':
    main()
