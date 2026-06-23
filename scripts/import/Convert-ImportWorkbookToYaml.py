#!/usr/bin/env python3
"""
Zentraler Excel-Import: Arbeitsbuch → YAML + Rückmeldung in import_status.

Nur Zeilen mit review_status=approved und import_status=offen|leer werden verarbeitet.
Nach erfolgreichem Export werden Zeilen auf import_status=übernommen gesetzt.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE = REPO_ROOT / 'import' / 'templates' / 'private-access-import-template.xlsx'
DEFAULT_INPUT = REPO_ROOT / 'import' / 'excel' / 'private-access-import.xlsx'
DEFAULT_OUTPUT = REPO_ROOT / 'config' / 'applications'
DEFAULT_MANIFEST = REPO_ROOT / 'import' / 'output' / 'import-manifest.json'

DEFAULT_OWNERS = [
    'florian.ried@extern.edeka.de',
    'ertan.simsek@extern.edeka.de',
]

OPEN_IMPORT_STATUSES = frozenset({'', 'offen', 'pending', 'open'})
SKIP_IMPORT_STATUSES = frozenset({'übernommen', 'generated', 'imported', 'übersprungen', 'skipped', 'fehler', 'error'})

HEADER_ALIASES = {
    'application_name': ('application_name', 'applicaton_name', 'app_name'),
    'connector_group': ('connector_group',),
    'entra_group_name': ('entra_group_name', 'entra_group', 'principal_name'),
    'target_type': ('target_type',),
    'fqdn': ('fqdn',),
    'ip_address': ('ip_address',),
    'cidr_range': ('cidr_range',),
    'ip_range_start': ('ip_range_start',),
    'ip_range_end': ('ip_range_end',),
    'protocol': ('protocol',),
    'ports': ('ports', 'port'),
    'port_range_end': ('port_range_end',),
    'change_reference': ('change_reference', 'change_reference_id'),
    'workload': ('workload', 'workload_tag'),
    'source_reference': ('source_reference',),
    'review_status': ('review_status',),
    'review_comment': ('review_comment',),
    'import_status': ('import_status',),
    'generated_yaml': ('generated_yaml',),
    'imported_at': ('imported_at',),
    'import_note': ('import_note',),
}


def require_openpyxl():
    if load_workbook is None:
        raise SystemExit(
            'openpyxl fehlt. Installieren: pip install -r import/requirements.txt'
        )


def normalize_header(value: str) -> str:
    return re.sub(r'[\s-]+', '_', (value or '').strip().lower())


def build_header_map(raw_headers: list[str]) -> dict[str, int]:
    normalized = [normalize_header(h) for h in raw_headers]
    mapping: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = normalize_header(alias)
            if alias_norm in normalized:
                mapping[canonical] = normalized.index(alias_norm)
                break
    return mapping


def cell_value(row: list, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ''
    value = row[idx]
    if value is None:
        return ''
    return str(value).strip()


def forward_fill(rows: list[dict]) -> None:
    last = {
        'application_name': '',
        'connector_group': '',
        'entra_group_name': '',
        'change_reference': '',
        'workload': '',
    }
    for rec in rows:
        for key in last:
            if rec.get(key):
                last[key] = rec[key]
            else:
                rec[key] = last[key]


def parse_ports(rec: dict) -> list[str]:
    ports_raw = rec.get('ports', '')
    port_end = rec.get('port_range_end', '')
    if ports_raw and port_end and '-' not in ports_raw:
        if ports_raw == port_end:
            return [ports_raw]
        return [f'{ports_raw}-{port_end}']

    if not ports_raw:
        return []

    result: list[str] = []
    for part in re.split(r'[,;]', ports_raw):
        part = part.strip()
        if part:
            result.append(part)
    return result


def host_for(rec: dict) -> str:
    tt = rec.get('target_type', '')
    if tt == 'fqdn':
        return rec.get('fqdn', '')
    if tt == 'ipAddress':
        return rec.get('ip_address', '')
    if tt == 'ipRangeCidr':
        return rec.get('cidr_range', '')
    if tt == 'ipRange':
        start = rec.get('ip_range_start', '')
        end = rec.get('ip_range_end', '')
        if start and end:
            return f'{start}-{end}'
        return start
    if tt == 'dnsSuffix':
        return rec.get('fqdn', '')
    return ''


def protocols_for(rec: dict) -> list[str]:
    p = (rec.get('protocol') or 'tcp').strip().lower().replace(' ', '')
    if p in ('both', 'tcp,udp', 'tcp+udp', 'tcp&udp', 'tcpudp'):
        return ['tcp,udp']
    if p == 'udp':
        return ['udp']
    return ['tcp']


def yaml_quote(value: str) -> str:
    if not value:
        return '""'
    if any(c in value for c in ':,#{}[]&*!|>\'"@`') or value[0] in '-?':
        return '"' + value.replace('\\', '\\\\').replace('"', '\\"') + '"'
    return value


def build_yaml(app_name: str, connector: str, destinations: list[dict], pilot_group: str,
               change_reference: str, workload: str) -> str:
    workload = workload or 'private-access'
    desc = f'Private Access – {app_name} ({connector}).'
    lines = [
        'apiVersion: gsa.gitops/v1',
        'kind: PrivateAccessApplication',
        'metadata:',
        f'  name: {yaml_quote(app_name)}',
        f'  description: {yaml_quote(desc)}',
        '  owners:',
    ]
    for owner in DEFAULT_OWNERS:
        lines.append(f'    - {owner}')
    lines += [
        f'  changeReference: {yaml_quote(change_reference or "IMPORT-001")}',
        '  tags:',
        f'    workload: {yaml_quote(workload)}',
        '    source: excel-import',
        'spec:',
        '  applicationType: enterprise',
        '  isAccessibleViaZTNAClient: true',
        f'  connectorGroup: {yaml_quote(connector)}',
        '  destinations:',
    ]
    for dest in destinations:
        lines.append(f'    - host: {yaml_quote(dest["host"])}')
        lines.append(f'      type: {dest["type"]}')
        ports_yaml = ', '.join(f'"{p}"' for p in dest['ports'])
        lines.append(f'      ports: [{ports_yaml}]')
        lines.append(f'      protocol: {dest["protocol"]}')
    lines += [
        '  assignments:',
        '    - principalType: Group',
        f'      principalName: {yaml_quote(pilot_group)}',
    ]
    return '\n'.join(lines) + '\n'


def import_status_open(value: str) -> bool:
    return normalize_header(value) in OPEN_IMPORT_STATUSES or value.strip() == ''


def review_approved(value: str) -> bool:
    return normalize_header(value) == 'approved'


def read_workbook(path: Path) -> tuple[list[str], list[list]]:
    require_openpyxl()
    wb = load_workbook(path)
    ws = wb['Import'] if 'Import' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    data = []
    for raw in rows[1:]:
        if not any(raw):
            continue
        data.append([str(c).strip() if c is not None else '' for c in raw])
    return headers, data


def read_csv(path: Path) -> tuple[list[str], list[list]]:
    with path.open(encoding='utf-8-sig', newline='') as handle:
        reader = csv.reader(handle, delimiter=';')
        rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def write_workbook(path: Path, headers: list[str], data_rows: list[list]) -> None:
    require_openpyxl()
    wb = load_workbook(path)
    ws = wb['Import'] if 'Import' in wb.sheetnames else wb.active
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    for row in data_rows:
        padded = row + [''] * max(0, len(headers) - len(row))
        ws.append(padded[:len(headers)])
    wb.save(path)


def rows_to_records(headers: list[str], data_rows: list[list]) -> list[dict]:
    header_map = build_header_map(headers)
    records: list[dict] = []
    for row_idx, row in enumerate(data_rows, start=2):
        rec = {'_row_index': row_idx, '_raw_row': row}
        for canonical, idx in header_map.items():
            rec[canonical] = cell_value(row, idx)
        if any(rec.get(k) for k in ['fqdn', 'ip_address', 'cidr_range', 'ip_range_start']):
            records.append(rec)
    forward_fill(records)
    return records


def records_to_sheet_rows(headers: list[str], records: list[dict], all_rows: list[list]) -> list[list]:
    header_map = build_header_map(headers)
    row_by_index = {rec['_row_index']: rec for rec in records}
    updated: list[list] = []
    for row_idx, row in enumerate(all_rows, start=2):
        rec = row_by_index.get(row_idx)
        new_row = list(row)
        while len(new_row) < len(headers):
            new_row.append('')
        if rec:
            for canonical, col_idx in header_map.items():
                if canonical in rec and col_idx is not None:
                    while len(new_row) <= col_idx:
                        new_row.append('')
                    new_row[col_idx] = rec.get(canonical, '')
        updated.append(new_row)
    return updated


def ensure_status_columns(headers: list[str]) -> list[str]:
    required = ['import_status', 'generated_yaml', 'imported_at', 'import_note']
    result = list(headers)
    for col in required:
        if normalize_header(col) not in {normalize_header(h) for h in result}:
            result.append(col)
    return result


def run(args: argparse.Namespace) -> int:
    input_path = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    manifest_path = Path(args.manifest).resolve()

    if not input_path.exists():
        raise SystemExit(
            f'Arbeitsbuch nicht gefunden: {input_path}\n'
            f'Kopieren Sie die Vorlage:\n'
            f'  cp import/templates/private-access-import-template.xlsx import/excel/private-access-import.xlsx'
        )

    suffix = input_path.suffix.lower()
    if suffix == '.csv':
        headers, data_rows = read_csv(input_path)
    elif suffix in ('.xlsx', '.xlsm'):
        headers, data_rows = read_workbook(input_path)
    else:
        raise SystemExit('Unterstützt: .xlsx, .xlsm, .csv')

    headers = ensure_status_columns(headers)
    records = rows_to_records(headers, data_rows)

    eligible: list[dict] = []
    for rec in records:
        status = rec.get('import_status', '')
        review = rec.get('review_status', '')
        if not review_approved(review):
            if args.mark_skipped and not status:
                rec['import_status'] = 'übersprungen'
                rec['import_note'] = f'review_status={review or "leer"}'
            continue
        if args.regenerate or import_status_open(status):
            eligible.append(rec)
        elif args.mark_skipped and status in SKIP_IMPORT_STATUSES:
            continue

    grouped: dict[tuple, list[dict]] = defaultdict(list)
    for rec in eligible:
        key = (
            rec.get('application_name', ''),
            rec.get('connector_group', ''),
            rec.get('entra_group_name', ''),
            rec.get('change_reference', '') or args.default_change_reference,
            rec.get('workload', '') or args.default_workload,
        )
        grouped[key].append(rec)

    manifest = []
    warnings: list[str] = []
    now = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')

    for key, rows in sorted(grouped.items()):
        app_name, connector, pilot_group, change_ref, workload = key
        if not app_name or not connector or not pilot_group:
            note = 'application_name, connector_group oder entra_group_name fehlt'
            for rec in rows:
                rec['import_status'] = 'fehler'
                rec['import_note'] = note
            warnings.append(f'{note}: Zeilen {[r["_row_index"] for r in rows]}')
            continue

        dest_map: dict[tuple, dict] = {}
        for rec in rows:
            host = host_for(rec)
            ports = parse_ports(rec)
            if not host or not ports:
                rec['import_status'] = 'fehler'
                rec['import_note'] = 'Ziel oder ports unvollständig'
                warnings.append(f'Zeile {rec["_row_index"]}: unvollständig')
                continue
            target_type = rec.get('target_type', '')
            for proto in protocols_for(rec):
                sig = (host.lower(), target_type, proto, tuple(ports))
                dest_map[sig] = {
                    'host': host,
                    'type': target_type,
                    'ports': ports,
                    'protocol': proto,
                }

        destinations = sorted(dest_map.values(), key=lambda d: (d['type'], d['host'], d['protocol']))
        if not destinations:
            warnings.append(f'Keine gültigen Segmente für {app_name}')
            continue
        if len(destinations) > 500:
            warnings.append(f'WARN >500 Segmente ({len(destinations)}): {app_name}')

        fname = f'{app_name}.yaml'
        content = build_yaml(app_name, connector, destinations, pilot_group, change_ref, workload)
        out_path = output_dir / fname

        if args.dry_run:
            print(f'[dry-run] würde schreiben: {out_path} ({len(destinations)} Segmente, {len(rows)} Zeilen)')
        else:
            output_dir.mkdir(parents=True, exist_ok=True)
            out_path.write_text(content, encoding='utf-8')
            print(f'geschrieben: {out_path.name} ({len(destinations)} Segmente, {len(rows)} Zeilen)')

        for rec in rows:
            if rec.get('import_status') == 'fehler':
                continue
            rec['import_status'] = 'übernommen'
            rec['generated_yaml'] = fname
            rec['imported_at'] = now
            rec['import_note'] = f'{len(destinations)} Segmente'

        manifest.append({
            'file': str(out_path.relative_to(REPO_ROOT)),
            'application': app_name,
            'connector_group': connector,
            'segments': len(destinations),
            'source_rows': len(rows),
            'row_numbers': [r['_row_index'] for r in rows],
        })

    if not args.dry_run and suffix in ('.xlsx', '.xlsm'):
        updated_rows = records_to_sheet_rows(headers, records, data_rows)
        write_workbook(input_path, headers, updated_rows)
        print(f'Arbeitsbuch aktualisiert: {input_path}')
    elif not args.dry_run and suffix == '.csv':
        warnings.append('CSV: import_status wird nicht zurückgeschrieben — bitte .xlsx nutzen.')

    if not args.dry_run:
        manifest_path.parent.mkdir(parents=True, exist_ok=True)
        manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False) + '\n', encoding='utf-8')
        print(f'Manifest: {manifest_path}')

    if warnings:
        print('\nHinweise:')
        for w in warnings[:30]:
            print(' ', w)

    print(f'\nApps: {len(manifest)}, verarbeitete Zeilen: {len(eligible)}')
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description='Excel-Arbeitsbuch → Private-Access-YAML')
    parser.add_argument(
        '-i', '--input',
        default=str(DEFAULT_INPUT),
        help=f'Arbeitsbuch (Standard: {DEFAULT_INPUT.relative_to(REPO_ROOT)})',
    )
    parser.add_argument(
        '-o', '--output',
        default=str(DEFAULT_OUTPUT),
        help='Zielordner für YAML-Dateien',
    )
    parser.add_argument(
        '--manifest',
        default=str(DEFAULT_MANIFEST),
        help='JSON-Manifest des Laufs',
    )
    parser.add_argument('--dry-run', action='store_true', help='Nur anzeigen, nichts schreiben')
    parser.add_argument('--regenerate', action='store_true', help='Auch bereits übernommene Zeilen erneut verarbeiten')
    parser.add_argument('--mark-skipped', action='store_true', help='Nicht-approved Zeilen als übersprungen markieren')
    parser.add_argument('--default-change-reference', default='IMPORT-001')
    parser.add_argument('--default-workload', default='private-access')
    args = parser.parse_args()
    return run(args)


if __name__ == '__main__':
    sys.exit(main())
